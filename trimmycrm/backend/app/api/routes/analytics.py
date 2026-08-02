from __future__ import annotations

import csv
import io
from datetime import UTC, datetime, timedelta
from decimal import Decimal
from uuid import UUID

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill
from sqlalchemy import and_, case, func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.deps import actor_tenant_db, actor_tenant_id, require_owner
from app.core.errors import BadRequestError
from app.models import (
    Appointment,
    AppointmentStatus,
    Pet,
    PlatformUser,
    Service,
    Site,
    Staff,
    TenantUser,
)
from app.services.access import plan_access_for_tenant
from app.services.scheduling import parse_timezone, working_ranges_for_date

router = APIRouter(tags=["analytics"])


def _period(from_: datetime, to: datetime) -> tuple[datetime, datetime]:
    if from_.tzinfo is None or to.tzinfo is None:
        raise BadRequestError("Границы периода должны содержать timezone", code="timezone_required")
    if to <= from_ or to - from_ > timedelta(days=366):
        raise BadRequestError("Некорректный диапазон", code="invalid_date_range")
    return from_.astimezone(UTC), to.astimezone(UTC)


@router.get("/analytics/overview")
async def overview(
    from_: datetime = Query(alias="from"),
    to: datetime = Query(),
    _owner: PlatformUser = Depends(require_owner),
    tenant_id: UUID = Depends(actor_tenant_id),
    session: AsyncSession = Depends(actor_tenant_db, scope="function"),
) -> dict[str, object]:
    access = await plan_access_for_tenant(session, tenant_id)
    if "basic_analytics" not in access.features and "advanced_analytics" not in access.features:
        access.require("basic_analytics")
    start, end = _period(from_, to)
    appointment_filters = [
        Appointment.tenant_id == tenant_id,
        Appointment.start_at >= start,
        Appointment.start_at < end,
    ]
    count = int(
        await session.scalar(
            select(func.count())
            .select_from(Appointment)
            .where(
                *appointment_filters,
                Appointment.status != AppointmentStatus.cancelled,
            )
        )
        or 0
    )
    revenue = Decimal(
        await session.scalar(
            select(func.coalesce(func.sum(Appointment.price), 0)).where(
                *appointment_filters,
                Appointment.status == AppointmentStatus.completed,
            )
        )
        or 0
    )
    new_clients = int(
        await session.scalar(
            select(func.count())
            .select_from(TenantUser)
            .where(
                TenantUser.tenant_id == tenant_id,
                TenantUser.created_at >= start,
                TenantUser.created_at < end,
            )
        )
        or 0
    )
    occupied_seconds = float(
        await session.scalar(
            select(
                func.coalesce(
                    func.sum(func.extract("epoch", Appointment.end_at - Appointment.start_at)), 0
                )
            ).where(
                *appointment_filters,
                Appointment.status.in_(
                    [
                        AppointmentStatus.new,
                        AppointmentStatus.confirmed,
                        AppointmentStatus.completed,
                    ]
                ),
            )
        )
        or 0
    )
    site = await session.get(Site, tenant_id)
    staff_rows = (
        await session.scalars(
            select(Staff).where(Staff.tenant_id == tenant_id, Staff.is_active.is_(True))
        )
    ).all()
    available_seconds = 0.0
    if site:
        tz = parse_timezone(getattr(site, "timezone", "Europe/Moscow"))
        current_day = start.astimezone(tz).date()
        final_day = end.astimezone(tz).date()
        while current_day <= final_day:
            for staff in staff_rows:
                schedule = staff.schedule or site.work_hours
                for value in working_ranges_for_date(schedule, current_day, tz):
                    clipped_start = max(value.starts_at.astimezone(UTC), start)
                    clipped_end = min(value.ends_at.astimezone(UTC), end)
                    if clipped_end > clipped_start:
                        available_seconds += (clipped_end - clipped_start).total_seconds()
            current_day += timedelta(days=1)
    utilization = round(100 * occupied_seconds / available_seconds, 2) if available_seconds else 0.0
    return {
        "from": start,
        "to": end,
        "appointments": count,
        "revenue": revenue,
        "newClients": new_clients,
        "staffUtilizationPercent": min(utilization, 100.0),
        "formula": {
            "revenue": "sum(price) for completed appointments",
            "utilization": "occupied appointment minutes / scheduled staff minutes",
        },
    }


@router.get("/analytics/services")
async def service_analytics(
    from_: datetime | None = Query(default=None, alias="from"),
    to: datetime | None = None,
    _owner: PlatformUser = Depends(require_owner),
    tenant_id: UUID = Depends(actor_tenant_id),
    session: AsyncSession = Depends(actor_tenant_db, scope="function"),
) -> list[dict[str, object]]:
    access = await plan_access_for_tenant(session, tenant_id)
    if "basic_analytics" not in access.features and "advanced_analytics" not in access.features:
        access.require("basic_analytics")
    end = (to or datetime.now(UTC)).astimezone(UTC)
    start = (from_ or end - timedelta(days=30)).astimezone(UTC)
    _period(start, end)
    rows = (
        await session.execute(
            select(
                Service.id,
                Service.name,
                func.count(Appointment.id),
                func.coalesce(
                    func.sum(
                        case(
                            (Appointment.status == AppointmentStatus.completed, Appointment.price),
                            else_=0,
                        )
                    ),
                    0,
                ),
            )
            .join(
                Appointment,
                and_(
                    Appointment.tenant_id == Service.tenant_id,
                    Appointment.service_id == Service.id,
                    Appointment.start_at >= start,
                    Appointment.start_at < end,
                ),
                isouter=True,
            )
            .where(Service.tenant_id == tenant_id)
            .group_by(Service.id, Service.name)
            .order_by(func.count(Appointment.id).desc())
        )
    ).all()
    return [
        {"serviceId": row[0], "serviceName": row[1], "appointments": row[2], "revenue": row[3]}
        for row in rows
    ]


async def _dashboard_export_data(
    *,
    from_: datetime,
    to: datetime,
    owner: PlatformUser,
    tenant_id: UUID,
    session: AsyncSession,
) -> tuple[dict[str, object], list[dict[str, object]]]:
    (await plan_access_for_tenant(session, tenant_id)).require("export")
    metrics = await overview(
        from_=from_,
        to=to,
        _owner=owner,
        tenant_id=tenant_id,
        session=session,
    )
    services = await service_analytics(
        from_=from_,
        to=to,
        _owner=owner,
        tenant_id=tenant_id,
        session=session,
    )
    return metrics, services


def _dashboard_xlsx_response(
    filename: str,
    metrics: dict[str, object],
    services: list[dict[str, object]],
) -> StreamingResponse:
    workbook = Workbook()
    overview_sheet = workbook.active
    if overview_sheet is None:
        raise RuntimeError("Новая книга Excel не содержит активного листа")
    overview_sheet.title = "Обзор"
    overview_sheet.append(["Показатель", "Значение"])
    overview_sheet.append(["Начало периода", _safe_cell(metrics["from"])])
    overview_sheet.append(["Конец периода", _safe_cell(metrics["to"])])
    overview_sheet.append(["Записи", _safe_cell(metrics["appointments"])])
    overview_sheet.append(["Выручка", _safe_cell(metrics["revenue"])])
    overview_sheet.append(["Новые клиенты", _safe_cell(metrics["newClients"])])
    overview_sheet.append(["Загрузка команды, %", _safe_cell(metrics["staffUtilizationPercent"])])

    service_sheet = workbook.create_sheet("Услуги")
    service_sheet.append(["Услуга", "Записи", "Выручка"])
    for service in services:
        service_sheet.append(
            [
                _safe_cell(service["serviceName"]),
                _safe_cell(service["appointments"]),
                _safe_cell(service["revenue"]),
            ]
        )

    header_fill = PatternFill("solid", fgColor="D8FF3E")
    for sheet in (overview_sheet, service_sheet):
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
        sheet.freeze_panes = "A2"
        sheet.column_dimensions["A"].width = 34
        sheet.column_dimensions["B"].width = 20
    service_sheet.column_dimensions["C"].width = 20

    if services:
        chart = BarChart()
        chart.type = "bar"
        chart.style = 10
        chart.title = "Популярность услуг"
        chart.y_axis.title = "Услуга"
        chart.x_axis.title = "Количество записей"
        data = Reference(service_sheet, min_col=2, min_row=1, max_row=len(services) + 1)
        categories = Reference(service_sheet, min_col=1, min_row=2, max_row=len(services) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.height = max(7, min(16, len(services) * 0.65))
        chart.width = 18
        service_sheet.add_chart(chart, "E2")

    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _safe_cell(value: object) -> object:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, UUID):
        return str(value)
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


def _csv_response(filename: str, headers: list[str], rows: list[list[object]]) -> StreamingResponse:
    stream = io.StringIO(newline="")
    stream.write("\ufeff")
    writer = csv.writer(stream, delimiter=";")
    writer.writerow(headers)
    writer.writerows([[_safe_cell(value) for value in row] for row in rows])
    content = stream.getvalue().encode("utf-8")
    return StreamingResponse(
        iter([content]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _xlsx_response(
    filename: str, sheet_name: str, headers: list[str], rows: list[list[object]]
) -> StreamingResponse:
    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet(sheet_name)
    sheet.append(headers)
    for row in rows:
        sheet.append([_safe_cell(value) for value in row])
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


async def _client_export_rows(session: AsyncSession, tenant_id: UUID) -> list[list[object]]:
    rows = (
        await session.execute(
            select(
                TenantUser.id,
                TenantUser.full_name,
                TenantUser.email,
                TenantUser.phone,
                func.count(Pet.id),
                TenantUser.created_at,
            )
            .outerjoin(
                Pet, and_(Pet.tenant_id == TenantUser.tenant_id, Pet.owner_id == TenantUser.id)
            )
            .where(TenantUser.tenant_id == tenant_id)
            .group_by(TenantUser.id)
            .order_by(TenantUser.created_at)
        )
    ).all()
    return [list(row) for row in rows]


async def _appointment_export_rows(session: AsyncSession, tenant_id: UUID) -> list[list[object]]:
    rows = (
        await session.execute(
            select(
                Appointment.id,
                Appointment.start_at,
                Appointment.end_at,
                TenantUser.full_name,
                Service.name,
                Staff.name,
                Appointment.status,
                Appointment.price,
                Appointment.prepaid,
            )
            .join(
                TenantUser,
                and_(
                    TenantUser.tenant_id == Appointment.tenant_id,
                    TenantUser.id == Appointment.tenant_user_id,
                ),
            )
            .join(
                Service,
                and_(
                    Service.tenant_id == Appointment.tenant_id, Service.id == Appointment.service_id
                ),
            )
            .outerjoin(
                Staff,
                and_(Staff.tenant_id == Appointment.tenant_id, Staff.id == Appointment.staff_id),
            )
            .where(Appointment.tenant_id == tenant_id)
            .order_by(Appointment.start_at)
        )
    ).all()
    return [list(row) for row in rows]


CLIENT_HEADERS = ["id", "name", "email", "phone", "pets", "created_at"]
APPOINTMENT_HEADERS = [
    "id",
    "start_at",
    "end_at",
    "client",
    "service",
    "staff",
    "status",
    "price",
    "prepaid",
]


@router.get("/analytics/export/dashboard.csv")
async def export_dashboard_csv(
    from_: datetime = Query(alias="from"),
    to: datetime = Query(),
    owner: PlatformUser = Depends(require_owner),
    tenant_id: UUID = Depends(actor_tenant_id),
    session: AsyncSession = Depends(actor_tenant_db, scope="function"),
) -> StreamingResponse:
    metrics, services = await _dashboard_export_data(
        from_=from_, to=to, owner=owner, tenant_id=tenant_id, session=session
    )
    rows: list[list[object]] = [
        ["Период", "Начало", metrics["from"], ""],
        ["Период", "Конец", metrics["to"], ""],
        ["Показатель", "Записи", metrics["appointments"], ""],
        ["Показатель", "Выручка", metrics["revenue"], ""],
        ["Показатель", "Новые клиенты", metrics["newClients"], ""],
        ["Показатель", "Загрузка команды, %", metrics["staffUtilizationPercent"], ""],
    ]
    rows.extend(
        ["Услуга", service["serviceName"], service["appointments"], service["revenue"]]
        for service in services
    )
    return _csv_response(
        "analytics-dashboard.csv",
        ["Раздел", "Название", "Записи/значение", "Выручка"],
        rows,
    )


@router.get("/analytics/export/dashboard.xlsx")
async def export_dashboard_xlsx(
    from_: datetime = Query(alias="from"),
    to: datetime = Query(),
    owner: PlatformUser = Depends(require_owner),
    tenant_id: UUID = Depends(actor_tenant_id),
    session: AsyncSession = Depends(actor_tenant_db, scope="function"),
) -> StreamingResponse:
    metrics, services = await _dashboard_export_data(
        from_=from_, to=to, owner=owner, tenant_id=tenant_id, session=session
    )
    return _dashboard_xlsx_response("analytics-dashboard.xlsx", metrics, services)


@router.get("/export/clients.csv")
async def export_clients_csv(
    _owner: PlatformUser = Depends(require_owner),
    tenant_id: UUID = Depends(actor_tenant_id),
    session: AsyncSession = Depends(actor_tenant_db, scope="function"),
) -> StreamingResponse:
    (await plan_access_for_tenant(session, tenant_id)).require("export")
    return _csv_response(
        "clients.csv", CLIENT_HEADERS, await _client_export_rows(session, tenant_id)
    )


@router.get("/export/appointments.csv")
async def export_appointments_csv(
    _owner: PlatformUser = Depends(require_owner),
    tenant_id: UUID = Depends(actor_tenant_id),
    session: AsyncSession = Depends(actor_tenant_db, scope="function"),
) -> StreamingResponse:
    (await plan_access_for_tenant(session, tenant_id)).require("export")
    return _csv_response(
        "appointments.csv", APPOINTMENT_HEADERS, await _appointment_export_rows(session, tenant_id)
    )


@router.get("/export/clients.xlsx")
async def export_clients_xlsx(
    _owner: PlatformUser = Depends(require_owner),
    tenant_id: UUID = Depends(actor_tenant_id),
    session: AsyncSession = Depends(actor_tenant_db, scope="function"),
) -> StreamingResponse:
    (await plan_access_for_tenant(session, tenant_id)).require("export")
    return _xlsx_response(
        "clients.xlsx", "Clients", CLIENT_HEADERS, await _client_export_rows(session, tenant_id)
    )


@router.get("/export/appointments.xlsx")
async def export_appointments_xlsx(
    _owner: PlatformUser = Depends(require_owner),
    tenant_id: UUID = Depends(actor_tenant_id),
    session: AsyncSession = Depends(actor_tenant_db, scope="function"),
) -> StreamingResponse:
    (await plan_access_for_tenant(session, tenant_id)).require("export")
    return _xlsx_response(
        "appointments.xlsx",
        "Appointments",
        APPOINTMENT_HEADERS,
        await _appointment_export_rows(session, tenant_id),
    )
