from __future__ import annotations

import io
import os
from datetime import UTC, datetime
from decimal import Decimal
from uuid import UUID

import pytest
from openpyxl import load_workbook
from starlette.responses import StreamingResponse

for key, value in {
    "DATABASE_URL": "postgresql+asyncpg://trimmycrm:trimmycrm@localhost/trimmycrm_test",
    "REDIS_URL": "redis://localhost:6379/0",
    "CELERY_BROKER_URL": "redis://localhost:6379/1",
    "CELERY_RESULT_BACKEND": "redis://localhost:6379/2",
    "S3_ENDPOINT_URL": "https://storage.yandexcloud.net",
    "INTERNAL_EDGE_TOKEN": "edge-token-1111111111111111111111111111",
    "JWT_PLATFORM_SECRET": "platform-secret-111111111111111111111111",
    "JWT_TENANT_SECRET": "tenant-secret-22222222222222222222222222",
    "AUTH_TOKEN_PEPPER": "token-pepper-333333333333333333333333333",
    "ENVIRONMENT": "test",
    "PAYMENT_PROVIDER": "mock",
}.items():
    os.environ[key] = value

from app.api.routes.analytics import (  # noqa: E402
    APPOINTMENT_HEADERS,
    CLIENT_HEADERS,
    _csv_response,
    _dashboard_xlsx_response,
    _safe_cell,
    _xlsx_response,
)
from app.models import AppointmentStatus  # noqa: E402

ROW_ID = UUID("aaaaaaaa-aaaa-4aaa-8aaa-aaaaaaaaaaaa")
CREATED_AT = datetime(2026, 7, 15, 12, 30, tzinfo=UTC)


async def _response_body(response: StreamingResponse) -> bytes:
    chunks = [chunk async for chunk in response.body_iterator]
    return b"".join(chunk if isinstance(chunk, bytes) else chunk.encode() for chunk in chunks)


def test_safe_cell_serializes_uuid() -> None:
    assert _safe_cell(ROW_ID) == "aaaaaaaa-aaaa-4aaa-8aaa-aaaaaaaaaaaa"


@pytest.mark.asyncio
@pytest.mark.parametrize(
    ("sheet_name", "headers", "row", "expected"),
    [
        (
            "Clients",
            CLIENT_HEADERS,
            [ROW_ID, "Анна", "anna@example.com", "+79990000000", 2, CREATED_AT],
            [
                str(ROW_ID),
                "Анна",
                "anna@example.com",
                "'+79990000000",
                2,
                CREATED_AT.isoformat(),
            ],
        ),
        (
            "Appointments",
            APPOINTMENT_HEADERS,
            [
                ROW_ID,
                CREATED_AT,
                datetime(2026, 7, 15, 14, 30, tzinfo=UTC),
                "Анна",
                "Боня",
                "Груминг",
                "Мария",
                AppointmentStatus.completed,
                Decimal("3500.00"),
                True,
            ],
            [
                str(ROW_ID),
                CREATED_AT.isoformat(),
                "2026-07-15T14:30:00+00:00",
                "Анна",
                "Боня",
                "Груминг",
                "Мария",
                "completed",
                "3500.00",
                True,
            ],
        ),
    ],
)
async def test_xlsx_response_serializes_export_rows(
    sheet_name: str,
    headers: list[str],
    row: list[object],
    expected: list[object],
) -> None:
    response = _xlsx_response("export.xlsx", sheet_name, headers, [row])

    workbook = load_workbook(io.BytesIO(await _response_body(response)), read_only=True)
    sheet = workbook[sheet_name]
    values = list(sheet.iter_rows(values_only=True))

    assert list(values[0]) == headers
    assert list(values[1]) == expected


@pytest.mark.asyncio
async def test_csv_response_format_is_unchanged_and_uuid_is_text() -> None:
    response = _csv_response(
        "clients.csv",
        ["id", "name", "created_at", "balance"],
        [[ROW_ID, '=HYPERLINK("bad")', CREATED_AT, Decimal("3500.00")]],
    )

    content = (await _response_body(response)).decode("utf-8")

    assert content == (
        "\ufeffid;name;created_at;balance\r\n"
        f'{ROW_ID};"\'=HYPERLINK(""bad"")";{CREATED_AT.isoformat()};3500.00\r\n'
    )


@pytest.mark.asyncio
async def test_dashboard_xlsx_contains_overview_services_and_chart() -> None:
    response = _dashboard_xlsx_response(
        "analytics-dashboard.xlsx",
        {
            "from": CREATED_AT,
            "to": datetime(2026, 7, 17, 12, 30, tzinfo=UTC),
            "appointments": 3,
            "revenue": Decimal("7250.00"),
            "newClients": 2,
            "staffUtilizationPercent": 58.5,
        },
        [
            {
                "serviceName": "Комплексный уход",
                "appointments": 3,
                "revenue": Decimal("7250.00"),
            }
        ],
    )

    workbook = load_workbook(io.BytesIO(await _response_body(response)))

    assert workbook.sheetnames == ["Обзор", "Услуги"]
    assert workbook["Обзор"]["B2"].value == CREATED_AT.isoformat()
    assert workbook["Услуги"]["A2"].value == "Комплексный уход"
    assert len(workbook["Услуги"]._charts) == 1
